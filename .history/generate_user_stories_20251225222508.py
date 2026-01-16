from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "User Stories"

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
user_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
admin_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
criteria_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 50

# Add headers
headers = ["ID", "User Story", "Role", "Acceptance Criteria"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 30

# User Stories and Acceptance Criteria based on code analysis
user_stories = [
    # AUTHENTICATION & REGISTRATION (User)
    {
        "id": "US-01",
        "role": "User",
        "story": "As a user, I want to register with my email and verify it so that I can create a secure account",
        "criteria": [
            "GIVEN the registration page is open",
            "WHEN I enter fullName, userName, phoneNumber, email, and password",
            "THEN I should receive an OTP on my email",
            "AND WHEN I verify the OTP within the time limit",
            "THEN my account should be created successfully"
        ]
    },
    {
        "id": "US-02",
        "role": "User",
        "story": "As a user, I want to send my email to receive an OTP so that I can verify my identity",
        "criteria": [
            "GIVEN I am on the email verification page",
            "WHEN I enter my email address",
            "THEN an OTP should be generated and sent to my email",
            "AND the OTP should be valid for a specific time period"
        ]
    },
    {
        "id": "US-03",
        "role": "User",
        "story": "As a user, I want to verify my email with OTP so that I can proceed with registration",
        "criteria": [
            "GIVEN I have received an OTP on my email",
            "WHEN I enter the correct OTP",
            "THEN the system should confirm verification",
            "AND I should see a success message"
        ]
    },
    {
        "id": "US-04",
        "role": "User",
        "story": "As a user, I want to login with my username or email and password so that I can access my account",
        "criteria": [
            "GIVEN I am on the login page",
            "WHEN I enter my username or email and correct password",
            "THEN I should be authenticated successfully",
            "AND I should receive access and refresh tokens",
            "AND I should be redirected to the home page"
        ]
    },
    {
        "id": "US-05",
        "role": "User",
        "story": "As a user, I want to logout from my account so that my session ends securely",
        "criteria": [
            "GIVEN I am logged in",
            "WHEN I click the logout button",
            "THEN my refresh token should be cleared",
            "AND I should be redirected to the login page"
        ]
    },
    {
        "id": "US-06",
        "role": "User",
        "story": "As a user, I want to reset my password if I forget it so that I can regain access",
        "criteria": [
            "GIVEN I click on forget password",
            "WHEN I enter my registered email",
            "THEN a password reset email with a reset link should be sent",
            "AND the reset token should be valid for 15 minutes"
        ]
    },
    {
        "id": "US-07",
        "role": "User",
        "story": "As a user, I want to set a new password using the reset link so that I can access my account again",
        "criteria": [
            "GIVEN I have a valid reset token",
            "WHEN I click the reset link and enter a new password",
            "THEN my password should be updated",
            "AND the reset token should be invalidated"
        ]
    },
    {
        "id": "US-08",
        "role": "User",
        "story": "As a user, I want to refresh my access token using the refresh token so that my session remains active",
        "criteria": [
            "GIVEN my access token has expired",
            "WHEN I send my refresh token",
            "THEN I should receive a new access token",
            "AND my session should continue without re-login"
        ]
    },

    # USER PROFILE (User)
    {
        "id": "US-09",
        "role": "User",
        "story": "As a user, I want to view my profile with all my details so that I can see my account information",
        "criteria": [
            "GIVEN I am logged in",
            "WHEN I navigate to my profile",
            "THEN I should see my fullName, userName, email, phoneNumber, location, gender, bio, and avatar",
            "AND no sensitive data like password should be visible"
        ]
    },
    {
        "id": "US-10",
        "role": "User",
        "story": "As a user, I want to update my profile information so that my account details are current",
        "criteria": [
            "GIVEN I am on my profile page",
            "WHEN I modify fullName, phoneNumber, location, gender, bio, or avatar",
            "THEN my profile should be updated successfully",
            "AND I should see a confirmation message"
        ]
    },

    # PRODUCT MANAGEMENT - ADMIN
    {
        "id": "US-11",
        "role": "Admin",
        "story": "As an admin, I want to add a new product with details so that I can manage inventory",
        "criteria": [
            "GIVEN I am on the add product page",
            "WHEN I enter name, description, price, stock (max 100), category, image, and expiry date",
            "THEN the product should be saved successfully",
            "AND I should receive a confirmation notification",
            "AND all admins should be notified about the new product"
        ]
    },
    {
        "id": "US-12",
        "role": "Admin",
        "story": "As an admin, I want to prevent duplicate products by name so that I don't create duplicates",
        "criteria": [
            "GIVEN I try to add a product",
            "WHEN the product name already exists for my admin ID",
            "THEN the system should show an error message",
            "AND the product should not be created"
        ]
    },
    {
        "id": "US-13",
        "role": "Admin",
        "story": "As an admin, I want to set a low stock threshold for products so that I get alerts",
        "criteria": [
            "GIVEN I am adding or editing a product",
            "WHEN I set a lowStockThreshold value",
            "THEN the product should store this threshold",
            "AND I should receive alerts when stock falls to or below this level"
        ]
    },
    {
        "id": "US-14",
        "role": "Admin",
        "story": "As an admin, I want to receive low stock alerts so that I can restock products in time",
        "criteria": [
            "GIVEN a product stock is at or below the low stock threshold",
            "WHEN the product is created or stock is updated",
            "THEN all admins should receive a notification",
            "AND the notification should show the product name and current stock"
        ]
    },
    {
        "id": "US-15",
        "role": "Admin",
        "story": "As an admin, I want to view all products with pagination so that I can manage them efficiently",
        "criteria": [
            "GIVEN I am on the product management page",
            "WHEN I view the products list",
            "THEN products should be displayed with pagination (default 4 per page)",
            "AND I should see current page and total pages information"
        ]
    },
    {
        "id": "US-16",
        "role": "Admin",
        "story": "As an admin, I want to search products by name or description so that I can find specific products",
        "criteria": [
            "GIVEN I am viewing products",
            "WHEN I enter a search term",
            "THEN only products matching the search criteria should be displayed",
            "AND the search should be case-insensitive"
        ]
    },
    {
        "id": "US-17",
        "role": "Admin",
        "story": "As an admin, I want to filter products by category so that I can view specific product types",
        "criteria": [
            "GIVEN I am viewing products",
            "WHEN I select a category filter",
            "THEN only products in that category should be displayed"
        ]
    },
    {
        "id": "US-18",
        "role": "Admin",
        "story": "As an admin, I want to filter products by availability status so that I can track out-of-stock items",
        "criteria": [
            "GIVEN I am viewing products",
            "WHEN I apply availability filter (available, low stock, or out of stock)",
            "THEN products should be filtered accordingly",
            "AND I should see the appropriate products for each filter"
        ]
    },
    {
        "id": "US-19",
        "role": "Admin",
        "story": "As an admin, I want to filter products by price range so that I can manage products by cost",
        "criteria": [
            "GIVEN I am viewing products",
            "WHEN I set minimum and maximum price",
            "THEN only products within the price range should be displayed"
        ]
    },
    {
        "id": "US-20",
        "role": "Admin",
        "story": "As an admin, I want to edit product details so that I can update information",
        "criteria": [
            "GIVEN I am viewing a product",
            "WHEN I click edit and modify name, description, price, or stock",
            "THEN the product should be updated successfully",
            "AND all admins should be notified about the update",
            "AND if stock falls below threshold, low stock alert should be sent"
        ]
    },
    {
        "id": "US-21",
        "role": "Admin",
        "story": "As an admin, I want to change product availability status so that I can enable or disable products",
        "criteria": [
            "GIVEN I am viewing a product",
            "WHEN I toggle the availability status",
            "THEN the product status should be updated",
            "AND all admins should be notified about the change"
        ]
    },
    {
        "id": "US-22",
        "role": "Admin",
        "story": "As an admin, I want to delete products from inventory so that I can remove obsolete items",
        "criteria": [
            "GIVEN I am viewing a product",
            "WHEN I click delete",
            "THEN the product should be removed from the database",
            "AND all admins should be notified about the deletion"
        ]
    },

    # PRODUCT BROWSING - USER
    {
        "id": "US-23",
        "role": "User",
        "story": "As a user, I want to browse available products so that I can see what's available for purchase",
        "criteria": [
            "GIVEN I am on the products page",
            "WHEN I load the page",
            "THEN all available products should be displayed with pagination",
            "AND I should see product name, price, image, and stock status"
        ]
    },
    {
        "id": "US-24",
        "role": "User",
        "story": "As a user, I want to search products by name or description so that I can find specific items",
        "criteria": [
            "GIVEN I am browsing products",
            "WHEN I enter a search term",
            "THEN only products matching the search should be displayed",
            "AND the search should be case-insensitive"
        ]
    },
    {
        "id": "US-25",
        "role": "User",
        "story": "As a user, I want to filter products by category so that I can view specific product types",
        "criteria": [
            "GIVEN I am browsing products",
            "WHEN I select a category",
            "THEN only products in that category should be displayed"
        ]
    },
    {
        "id": "US-26",
        "role": "User",
        "story": "As a user, I want to filter products by availability so that I can see in-stock items",
        "criteria": [
            "GIVEN I am browsing products",
            "WHEN I filter by availability status",
            "THEN only available products should be shown"
        ]
    },
    {
        "id": "US-27",
        "role": "User",
        "story": "As a user, I want to filter products by price range so that I can find affordable items",
        "criteria": [
            "GIVEN I am browsing products",
            "WHEN I set minimum and maximum price",
            "THEN only products within the range should be displayed"
        ]
    },
    {
        "id": "US-28",
        "role": "User",
        "story": "As a user, I want to view product details in a modal so that I can see complete information",
        "criteria": [
            "GIVEN I click on a product",
            "WHEN the product detail modal opens",
            "THEN I should see name, description, price, stock, image, category, and expiry date",
            "AND I should have an option to add to cart or buy"
        ]
    },

    # PURCHASING - USER
    {
        "id": "US-29",
        "role": "User",
        "story": "As a user, I want to buy products from the store so that I can make purchases",
        "criteria": [
            "GIVEN I have selected products to buy",
            "WHEN I confirm the purchase",
            "THEN the product stock should be reduced",
            "AND a purchase record should be created with 'pending' status",
            "AND I should receive a purchase confirmation notification"
        ]
    },
    {
        "id": "US-30",
        "role": "User",
        "story": "As a user, I want to cancel my order within 1 hour of purchase so that I can change my mind",
        "criteria": [
            "GIVEN I have placed an order less than 1 hour ago",
            "WHEN I click the cancel order button",
            "THEN the order should be marked as 'cancelled'",
            "AND the stock should be restored",
            "AND admins should be notified about the cancellation"
        ]
    },
    {
        "id": "US-31",
        "role": "User",
        "story": "As a user, I want to see that I cannot cancel orders after 1 hour so that I understand the policy",
        "criteria": [
            "GIVEN my order is older than 1 hour",
            "WHEN I try to cancel it",
            "THEN I should see an error message",
            "AND I should be asked to contact admin for cancellation"
        ]
    },
    {
        "id": "US-32",
        "role": "User",
        "story": "As a user, I want to buy using different payment gateways so that I have flexible payment options",
        "criteria": [
            "GIVEN I am checking out",
            "WHEN I select a payment method (Cash, Esewa, or Khalti)",
            "THEN the purchase record should store the payment gateway",
            "AND the transaction should proceed with the selected method"
        ]
    },
    {
        "id": "US-33",
        "role": "User",
        "story": "As a user, I want to create a payment with Esewa so that I can pay online securely",
        "criteria": [
            "GIVEN I select Esewa as payment method",
            "WHEN I enter the amount",
            "THEN a unique transaction UUID should be generated",
            "AND the payment form with signed fields should be prepared",
            "AND payment should be processed through Esewa"
        ]
    },

    # ORDER MANAGEMENT
    {
        "id": "US-34",
        "role": "User",
        "story": "As a user, I want to view my booked products/orders so that I can track my purchases",
        "criteria": [
            "GIVEN I am on my booked products page",
            "WHEN I load the page",
            "THEN all my orders should be displayed with pagination",
            "AND I should see order status, product name, quantity, and price",
            "AND orders should be sorted by newest first"
        ]
    },
    {
        "id": "US-35",
        "role": "User",
        "story": "As a user, I want to filter my orders by status so that I can track specific orders",
        "criteria": [
            "GIVEN I am viewing my orders",
            "WHEN I apply a status filter (pending, completed, cancelled)",
            "THEN only orders with that status should be displayed"
        ]
    },
    {
        "id": "US-36",
        "role": "User",
        "story": "As a user, I want to search my orders by product name so that I can find specific purchases",
        "criteria": [
            "GIVEN I am viewing my orders",
            "WHEN I search by product name",
            "THEN only matching orders should be displayed"
        ]
    },
    {
        "id": "US-37",
        "role": "Admin",
        "story": "As an admin, I want to view all customer orders so that I can manage fulfillment",
        "criteria": [
            "GIVEN I am on the manage booked products page",
            "WHEN I load the page",
            "THEN all orders from all users should be displayed with pagination",
            "AND I should see user details, product name, and order status",
            "AND orders should be sorted by newest first"
        ]
    },
    {
        "id": "US-38",
        "role": "Admin",
        "story": "As an admin, I want to filter all orders by status so that I can manage fulfillment workflow",
        "criteria": [
            "GIVEN I am viewing all orders",
            "WHEN I apply a status filter",
            "THEN only orders with that status should be displayed"
        ]
    },
    {
        "id": "US-39",
        "role": "Admin",
        "story": "As an admin, I want to search orders by username or customer name so that I can find customer orders",
        "criteria": [
            "GIVEN I am viewing all orders",
            "WHEN I search by username or fullName",
            "THEN only matching customer orders should be displayed"
        ]
    },
    {
        "id": "US-40",
        "role": "Admin",
        "story": "As an admin, I want to change order status so that I can manage order fulfillment",
        "criteria": [
            "GIVEN I am viewing an order",
            "WHEN I change the status to pending, completed, or cancelled",
            "THEN the order status should be updated",
            "AND the user should be notified of the status change",
            "AND all admins should be notified about the change"
        ]
    },
    {
        "id": "US-41",
        "role": "Admin",
        "story": "As an admin, I want stock to be restored when I cancel an order so that inventory is accurate",
        "criteria": [
            "GIVEN I am changing an order status from non-cancelled to cancelled",
            "WHEN I save the change",
            "THEN the product stock should be restored by the order quantity",
            "AND stock should not exceed the product's maximum capacity"
        ]
    },
    {
        "id": "US-42",
        "role": "Admin",
        "story": "As an admin, I want stock to be deducted again when I restore a cancelled order so that inventory is correct",
        "criteria": [
            "GIVEN I change an order status from cancelled to something else",
            "WHEN I save the change",
            "THEN the product stock should be deducted again",
            "AND the system should ensure sufficient stock is available"
        ]
    },

    # BILLING
    {
        "id": "US-43",
        "role": "User",
        "story": "As a user, I want to view my bill/invoice so that I can see transaction details",
        "criteria": [
            "GIVEN I click on generate bill",
            "WHEN the bill is generated",
            "THEN I should see product names, prices, quantities, and total amount",
            "AND I should see the payment gateway used"
        ]
    },
    {
        "id": "US-44",
        "role": "Admin",
        "story": "As an admin, I want to generate bills for customers so that I can provide invoices",
        "criteria": [
            "GIVEN I am viewing a customer's orders",
            "WHEN I generate their bill",
            "THEN a detailed invoice should be created",
            "AND I should see all their purchases, amounts, and payment methods"
        ]
    },

    # NOTIFICATIONS
    {
        "id": "US-45",
        "role": "User",
        "story": "As a user, I want to receive notifications about my purchases so that I stay informed",
        "criteria": [
            "GIVEN I perform an action (buy, cancel, profile update)",
            "WHEN the action is completed",
            "THEN a notification should be created in my notification list",
            "AND I should see the notification message with timestamp"
        ]
    },
    {
        "id": "US-46",
        "role": "User",
        "story": "As a user, I want to view my notifications with pagination so that I can check updates",
        "criteria": [
            "GIVEN I am on the notifications page",
            "WHEN I load the page",
            "THEN my notifications should be displayed with pagination (default 4 per page)",
            "AND unread notifications should be highlighted"
        ]
    },
    {
        "id": "US-47",
        "role": "User",
        "story": "As a user, I want to filter notifications by read/unread status so that I can prioritize",
        "criteria": [
            "GIVEN I am viewing notifications",
            "WHEN I apply a read/unread filter",
            "THEN only notifications with that status should be displayed"
        ]
    },
    {
        "id": "US-48",
        "role": "User",
        "story": "As a user, I want to mark notifications as read so that I can track what I've seen",
        "criteria": [
            "GIVEN I am viewing a notification",
            "WHEN I mark it as read",
            "THEN the notification status should be updated",
            "AND the notification should no longer appear as unread"
        ]
    },
    {
        "id": "US-49",
        "role": "Admin",
        "story": "As an admin, I want to receive notifications about inventory changes so that I stay updated",
        "criteria": [
            "GIVEN products are created, updated, deleted, or availability changes",
            "WHEN the action occurs",
            "THEN all admins should receive a notification",
            "AND the notification should show the action and product details"
        ]
    },
    {
        "id": "US-50",
        "role": "Admin",
        "story": "As an admin, I want to receive notifications about orders so that I can manage fulfillment",
        "criteria": [
            "GIVEN orders are placed, status changes, or user cancellations",
            "WHEN the action occurs",
            "THEN all admins should receive a notification",
            "AND the notification should show the order details"
        ]
    },

    # USER MANAGEMENT - ADMIN
    {
        "id": "US-51",
        "role": "Admin",
        "story": "As an admin, I want to view all registered users so that I can manage accounts",
        "criteria": [
            "GIVEN I am on the manage users page",
            "WHEN I load the page",
            "THEN all users should be displayed with pagination (default 4 per page)",
            "AND I should see user fullName, userName, email, and status"
        ]
    },
    {
        "id": "US-52",
        "role": "Admin",
        "story": "As an admin, I want to search users by name, username, or email so that I can find specific users",
        "criteria": [
            "GIVEN I am viewing all users",
            "WHEN I enter a search term",
            "THEN only matching users should be displayed",
            "AND search should be case-insensitive across multiple fields"
        ]
    },
    {
        "id": "US-53",
        "role": "Admin",
        "story": "As an admin, I want to activate or deactivate user accounts so that I can manage account access",
        "criteria": [
            "GIVEN I am viewing a user",
            "WHEN I toggle their active status",
            "THEN the user's isActive flag should be updated",
            "AND the user should be notified of the status change",
            "AND all admins should be notified about the change"
        ]
    },
    {
        "id": "US-54",
        "role": "Admin",
        "story": "As an admin, I want to change user roles so that I can promote or demote users",
        "criteria": [
            "GIVEN I am viewing a user",
            "WHEN I change their role from 'user' to 'admin' or vice versa",
            "THEN the role should be updated in the database",
            "AND the user should be notified of the role change",
            "AND all admins should be notified about the change"
        ]
    },
    {
        "id": "US-55",
        "role": "User",
        "story": "As a user, I want to see my login failure when I use wrong credentials so that I know why I failed",
        "criteria": [
            "GIVEN I am on the login page",
            "WHEN I enter incorrect password or non-existent account",
            "THEN I should see an appropriate error message",
            "AND I should not be logged in"
        ]
    },
    {
        "id": "US-56",
        "role": "User",
        "story": "As a user, I want to subscribe to newsletters so that I get updates about products",
        "criteria": [
            "GIVEN I am on the subscription section",
            "WHEN I enter my email and click subscribe",
            "THEN my email should be stored as a subscriber",
            "AND I should see a success message"
        ]
    },

    # STATISTICS & ANALYTICS
    {
        "id": "US-57",
        "role": "User",
        "story": "As a user, I want to view my purchase statistics so that I can track my spending",
        "criteria": [
            "GIVEN I am on my stats page",
            "WHEN the page loads",
            "THEN I should see total spent, completed orders, pending orders, and cancelled orders",
            "AND I should see paginated purchase history with product details"
        ]
    },
    {
        "id": "US-58",
        "role": "Admin",
        "story": "As an admin, I want to view system-wide dashboard statistics so that I can monitor business",
        "criteria": [
            "GIVEN I am on the admin dashboard",
            "WHEN the page loads",
            "THEN I should see total users, active users, total products, out-of-stock products, and low-stock products",
            "AND I should see total revenue, completed bookings, pending bookings, and cancelled bookings",
            "AND I should see recent bookings with pagination and category-wise product breakdown"
        ]
    },

    # SECURITY
    {
        "id": "US-59",
        "role": "User",
        "story": "As a user, I want deactivated accounts to not be able to login so that security is maintained",
        "criteria": [
            "GIVEN my account is deactivated by an admin",
            "WHEN I try to login",
            "THEN I should see an error message about account deactivation",
            "AND I should not be logged in",
            "AND I should be asked to contact the admin"
        ]
    },
    {
        "id": "US-60",
        "role": "Admin",
        "story": "As an admin, I want JWT authentication for all protected endpoints so that API security is maintained",
        "criteria": [
            "GIVEN protected endpoints are accessed",
            "WHEN I send requests without valid JWT token",
            "THEN the request should be rejected with 401 Unauthorized",
            "AND WHEN I send with valid token, the request should be processed"
        ]
    },
]

# Add data to worksheet
row = 2
for story in user_stories:
    ws.cell(row=row, column=1).value = story["id"]
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=1).border = border
    
    ws.cell(row=row, column=2).value = f"As a {story['role'].lower()} I want {story['story'].split('I want')[1].strip()}"
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=row, column=2).border = border
    if story["role"] == "Admin":
        ws.cell(row=row, column=2).fill = admin_fill
    else:
        ws.cell(row=row, column=2).fill = user_fill
    
    ws.cell(row=row, column=3).value = story["role"]
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=3).border = border
    if story["role"] == "Admin":
        ws.cell(row=row, column=3).fill = admin_fill
    else:
        ws.cell(row=row, column=3).fill = user_fill
    
    criteria_text = "\n".join(story["criteria"])
    ws.cell(row=row, column=4).value = criteria_text
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.cell(row=row, column=4).border = border
    ws.cell(row=row, column=4).fill = criteria_fill
    
    ws.row_dimensions[row].height = max(30, len(story["criteria"]) * 20)
    row += 1

# Adjust row heights for better visibility
ws.row_dimensions[1].height = 25

# Save the workbook
output_path = r"d:\lastyear\stock-zen\user_stories_acceptance_criteria.xlsx"
wb.save(output_path)
print(f"✓ User Stories and Acceptance Criteria created successfully!")
print(f"✓ Total Stories: {len(user_stories)}")
print(f"✓ File saved at: {output_path}")
print(f"\nBreakdown:")
print(f"  - User Stories: {len([s for s in user_stories if s['role'] == 'User'])}")
print(f"  - Admin Stories: {len([s for s in user_stories if s['role'] == 'Admin'])}")
